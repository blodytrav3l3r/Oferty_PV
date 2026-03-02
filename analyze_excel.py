import openpyxl
import json

wb = openpyxl.load_workbook(r'g:\GitHub\Oferty_PV\Impuls_13.01.xlsm', data_only=True)

result = {}

# Extract component data from DN sheets
for dn in ['DN1000', 'DN1200', 'DN1500', 'DN2000']:
    ws = wb[dn]
    components = {}
    
    # Row 1: indices, Row 2: prices, Row 3: weights
    # Columns P onwards (col 16+)
    indices = {}
    for cell in ws[1]:
        if cell.column >= 16 and cell.value:
            indices[cell.column] = str(cell.value)
    
    prices = {}
    for cell in ws[2]:
        if cell.column >= 16 and cell.value is not None:
            prices[cell.column] = cell.value
    
    weights = {}
    for cell in ws[3]:
        if cell.column >= 16 and cell.value is not None:
            try:
                weights[cell.column] = float(str(cell.value).replace(' ', ''))
            except:
                weights[cell.column] = 0
    
    # Row 13 has component names
    names = {}
    for cell in ws[13]:
        if cell.column >= 16 and cell.value:
            names[cell.column] = str(cell.value)
    
    # Row 14 has heights
    heights_row = {}
    for cell in ws[14]:
        if cell.column >= 16 and cell.value is not None:
            heights_row[cell.column] = cell.value
    
    # Build component list
    comp_list = []
    for col in sorted(indices.keys()):
        idx = indices.get(col, '')
        name = names.get(col, '')
        price = prices.get(col, 0)
        weight = weights.get(col, 0)
        height = heights_row.get(col, 0)
        if idx and idx != 'INDEKS' and price != 1000000 and 'BRAK' not in str(idx):
            comp_list.append({
                'col': col,
                'id': idx,
                'name': name,
                'price': price,
                'weight': weight,
                'height': height
            })
    
    components['nadbudowa'] = comp_list
    result[dn] = components

# Extract bottom part pricing from sheets 1000, 1200, 1500, 2000
for sheet_name, dn_key in [('1000', 'DN1000'), ('1200', 'DN1200'), ('1500', 'DN1500'), ('2000', 'DN2000')]:
    ws = wb[sheet_name]
    
    bottom = {}
    
    # Row 1: Cena (base prices) - columns E onwards for DN1000
    # Row 13: Wysokość robocza headers
    heights = []
    base_prices = []
    zelbet = []
    nierdzewka = []
    klasa_s = []
    wagi = []
    
    # Get heights from row 13 (starting from col D or E)
    for cell in ws[13]:
        if cell.column >= 4 and cell.value is not None:
            try:
                h = int(cell.value)
                heights.append({'col': cell.column, 'height': h})
            except:
                pass
    
    # Get base prices from row 1
    for h in heights:
        col = h['col']
        val = ws.cell(row=1, column=col).value
        base_prices.append(val if val else 0)
    
    # Get zelbet surcharge from row 2
    for h in heights:
        col = h['col']
        val = ws.cell(row=2, column=col).value
        zelbet.append(val if val else 0)
    
    # Get nierdzewka from row 3
    for h in heights:
        col = h['col']
        val = ws.cell(row=3, column=col).value
        nierdzewka.append(val if val else 0)
    
    # Get klasa S from row 4
    for h in heights:
        col = h['col']
        val = ws.cell(row=4, column=col).value
        klasa_s.append(val if val else 0)
    
    # Get weights from row 5
    for h in heights:
        col = h['col']
        val = ws.cell(row=5, column=col).value
        wagi.append(val if val else 0)
    
    # Get kineta prices from row 12 or cols AS/AT/AU in row 1
    kineta_prices = {}
    # Check AS, AT, AU columns from row 1
    as_col = None
    for cell in ws[1]:
        if cell.column >= 45 and cell.column <= 47 and cell.value is not None:
            kineta_prices[cell.column] = cell.value
    
    # Also look at row 12
    for cell in ws[12]:
        if cell.column >= 45 and cell.column <= 47 and cell.value is not None:
            kineta_prices[cell.column] = cell.value
    
    bottom['heights'] = [h['height'] for h in heights]
    bottom['basePrices'] = base_prices
    bottom['zelbetSurcharge'] = zelbet
    bottom['stainlessSurcharge'] = nierdzewka
    bottom['classSurcharge'] = klasa_s
    bottom['weights'] = wagi
    bottom['kineta'] = kineta_prices
    
    # Get index pattern from row 14
    index_pattern = []
    for h in heights:
        col = h['col']
        val = ws.cell(row=14, column=col).value
        index_pattern.append(str(val) if val else '')
    bottom['indexPattern'] = index_pattern
    
    # Transition prices from BB onwards in row 1 (Cena przed rabatem) and row 2 (Cena)
    # Row 13: material name, Row 14: diameters
    transitions = {}
    for cell in ws[13]:
        if cell.column >= 54 and cell.value and '???' not in str(cell.value):
            material = str(cell.value)
            diameter = ws.cell(row=14, column=cell.column).value
            price = ws.cell(row=2, column=cell.column).value
            if diameter and price and price != 10000:
                if material not in transitions:
                    transitions[material] = {}
                transitions[material][int(diameter)] = price
    
    bottom['transitions'] = transitions
    
    # UTH/PRE heights from rows 7-8 (AZ column area)
    uth_heights = {}
    pre_heights = {}
    for cell in ws[7]:
        if cell.column >= 54 and cell.value is not None:
            diameter = ws.cell(row=14, column=cell.column).value
            material = ws.cell(row=13, column=cell.column).value
            if diameter and str(cell.value) != 'BRAK':
                uth_heights[f"{material}_{diameter}"] = cell.value
    
    for cell in ws[8]:
        if cell.column >= 54 and cell.value is not None:
            diameter = ws.cell(row=14, column=cell.column).value
            material = ws.cell(row=13, column=cell.column).value
            if diameter and str(cell.value) != 'BRAK':
                pre_heights[f"{material}_{diameter}"] = cell.value
    
    result[dn_key]['bottom'] = bottom

# Write result
with open(r'g:\GitHub\Oferty_PV\extracted_studnie_data.json', 'w', encoding='utf-8') as f:
    json.dump(result, f, indent=2, ensure_ascii=False, default=str)

print("Done!")
